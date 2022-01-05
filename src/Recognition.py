import datetime
import os
import time

from openpyxl import load_workbook
import cv2
import pandas as pd
from numpy import asarray
import pickle
from keras.models import load_model
from numpy import asarray, load, expand_dims
from sklearn.preprocessing import LabelEncoder, Normalizer

from mtcnn.mtcnn import MTCNN
detector = MTCNN()
SVMModel = pickle.load(open("../SVMModel.sav", 'rb'))

FacenetModel = load_model('../facenet_keras.h5')

in_encoder = Normalizer(norm='l2')

data = load('../faceEmbedding-10samples.npz')
X, y= data['arr_0'],data['arr_1']

# X = in_encoder.transform(X)
out_encoder = LabelEncoder()
out_encoder.fit(y)
y = out_encoder.transform(y)

font = cv2.FONT_HERSHEY_SIMPLEX


fileName = ".."+os.sep+"StudentDetails"+os.sep+"StudentDetails.xlsx"
wb = load_workbook(fileName)
ws = wb.active
idCol = ws['A']
nameCol = ws['B']

Id = []
Name = []
for x in idCol[1:]: 
    # print(type(x.value))
    Id.append(x.value)
for x in nameCol[1:]: 
    Name.append(x.value)
# print (Id,Name)

ts = time.time()
date = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d')
cDate = ws.cell(row = 1, column=ws.max_column+1)
cDate.value=date

idAttendanced=[]

# print(Id)
def attendanceSuccess(n):
    id = n.split(".")[1]
    name = n.split(".")[0]
    if id not in idAttendanced:
        idAttendanced.append(id)
        ts = time.time()
        timeStamp = datetime.datetime.fromtimestamp(ts).strftime('%H:%M:%S')
        cTime = ws.cell(row = Id.index(int(id))+2, column=ws.max_column)
        cTime.value=timeStamp
    # print(idAttendanced)

def get_embedding(face_pixels):
	# scale pixel values
	face_pixels = face_pixels.astype('float32')
	# standardize pixel values across channels (global)
	mean, std = face_pixels.mean(), face_pixels.std()
	face_pixels = (face_pixels - mean) / std
	# transform face into one sample
	samples = expand_dims(face_pixels, axis=0)
	# make prediction to get embedding
	yhat = FacenetModel.predict(samples)
	return yhat[0]

def face_predict(face_embedded):
    sample = expand_dims(face_embedded, axis=0)
    sample = in_encoder.transform(sample)
    # print(sample.shape)
    yhat_class = SVMModel.predict(sample)
    yhat_prob = SVMModel.predict_proba(sample)
    class_index = yhat_class[0]
    class_probability = yhat_prob[0,class_index] * 100
    predict_name = out_encoder.inverse_transform(yhat_class)
    return predict_name[0], class_probability
#-------------------------
def recognizeAttendance():

    # Initialize and start realtime video capture
    cam = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    cam.set(3, 800)  # set video width
    cam.set(4, 640)  # set video height
    # Define min window size to be recognized as a face
    minW = 0.1 * cam.get(3)
    minH = 0.1 * cam.get(4)

    while(True):
        _,image = cam.read()
        pixels = asarray(image)
    
        results = detector.detect_faces(pixels)

        for f in results:
            x, y, w, h =  f['box']
            face = image[y:y+h, x:x+w]
            resize = cv2.resize(face, (160, 160))
            face_embedded = get_embedding(resize)
            predict_name, class_probability = face_predict(face_embedded)

            confstr = "{0}%".format(round(class_probability))
            if(class_probability > 70):
                cv2.putText(image, str(predict_name), (x+5,y-5), font, 0.5, (0, 255, 0), 2)
                cv2.putText(image, str(confstr), (x + 5, y + h - 5), font,0.5, (0, 255, 0), 2 )
                cv2.rectangle(image, (x, y), (x+w, y+h), (0, 255, 0), 2)
                attendanceSuccess(predict_name)

            elif (class_probability < 50):
                cv2.putText(image, str(predict_name), (x+5,y-5), font, 0.5, (0, 0, 255), 2)
                cv2.putText(image, str(confstr), (x + 5, y + h - 5), font,0.5, (0, 0, 255),2 )
                cv2.rectangle(image, (x, y), (x+w, y+h), (0, 0, 255), 2)
            else:
                cv2.putText(image, str(predict_name), (x+5,y-5), font, 0.5, (0, 165, 255), 2)
                cv2.putText(image, str(confstr), (x + 5, y + h - 5), font,0.5, (0, 165, 255),2 )
                cv2.rectangle(image, (x, y), (x+w, y+h), (0, 165, 255), 2)
        
        cv2.imshow('Attendance', image)
        if (cv2.waitKey(1) == ord('q')):
            break

    print("Attendance Successful")
    
    wb.save(fileName)
    wb.close()

    cam.release()
    cv2.destroyAllWindows()

# recognizeAttendance()