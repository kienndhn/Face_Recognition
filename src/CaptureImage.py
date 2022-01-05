import csv

import cv2
import os
from numpy import asarray
import os.path
from openpyxl import load_workbook
from mtcnn.mtcnn import MTCNN
detector = MTCNN()
# counting the numbers

id=[],
name=[]

fileName = ".."+os.sep+"StudentDetails"+os.sep+"StudentDetails.xlsx"
wb = load_workbook(fileName)
ws = wb.active
def readDetailFile():
    idCol = ws['A']
    nameCol = ws['B']
    Id = []
    Name = []
    for x in idCol[1:]: 
        # print(type(x.value))
        Id.append(x.value)
    for x in nameCol[1:]: 
        Name.append(x.value)
    # print (Id,Name)
    return Id, Name

def saveIdAndName(Id, name):
    cId = ws.cell(row = ws.max_row+1, column=1)
    cId.value=Id

    cName = ws.cell(row = ws.max_row, column=2)
    cName.value=name

    wb.save(fileName)
    wb.close()

# readDetailFile()

def checkIdExist(Id):
    id, name = readDetailFile()
    # print(type(Id))
    for i in id:
        if i == Id:
            return True
    # print (Id)
    return False

# checkIdExist(2)

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass

    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass

    return False



# Take image function

def takeImages():

    Id = input("Nhap ma sinh vien: ")
    name = input("Nhap ten: ")

    # id,_ = readDetailFile()

    if(name.isalpha()):
        if checkIdExist(Id) == False:
            
            cam = cv2.VideoCapture(0, cv2.CAP_DSHOW)
            sampleNum = 0

            directory = ".."+os.sep+"TrainingImage"+os.sep+name+"."+Id
            if not os.path.exists(directory):
                os.makedirs(directory)

            while(True):
                _, image = cam.read()
                pixels = asarray(image)
        
                results = detector.detect_faces(pixels)

                for f in results:
                    x, y, w, h =  f['box']
                    face = image[y:y+h, x:x+w]
                    resize = cv2.resize(face, (160, 160))
                    cv2.rectangle(image, (x, y), (x+w, y+h), (0, 255, 0), 2)
                    sampleNum = sampleNum+1
                    cv2.imwrite( directory+os.sep +
                                str(sampleNum) + ".jpg", resize)
                    cv2.imshow('Capture', image)

                if cv2.waitKey(100) & 0xFF == ord('q'):
                    break
                # break if the sample number is more than 100
                elif sampleNum == 100:
                    print("quit")
                    break
            cam.release()
            cv2.destroyAllWindows()
            res = "Da luu anh ID : " + Id + " Name : " + name
            saveIdAndName(Id, name)
        else:
            print("Id da ton tai")
            return takeImages()
    else:    
        print("Ten sai dinh dang")
        return takeImages()


